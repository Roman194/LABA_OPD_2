

from aiogram import Bot, Dispatcher, types, executor
from aiogram.dispatcher import filters
from openpyxl import load_workbook #импорт библиотеки для работы с таблицей

my_list = 'Registration_list.xlsx'
lwb = load_workbook(my_list)  # список листов файла
ws = lwb['Pivo']  # работа с конкретным листом файла
ws_1 = lwb['NoRule']
ws_2 = lwb['Classic']
ws_3 = lwb['Lgot']
ws_list=[ws,ws_1,ws_2,ws_3]
choosed=-1

buttons = ["1. Забег за пивом", "2. Забег без правил", "3. Классический", "4. Льготный"]
buttons_1 = ["Да", "Нет"]
bot = Bot(token="6207923861:AAEkzpberbx2cDAUZlDRGrbD7QRHAT9Nfkg")
dp = Dispatcher(bot)
user_data=[]


@dp.message_handler(commands=['start'])
async def choose_Race(message: types.Message):
    await message.reply("Привет! Это бот, который поможет тебе записаться на нужный марафон! Выбери один из забегов")
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(*buttons)

    await message.answer("Просто введи его номер:", reply_markup=keyboard)


@dp.message_handler(commands=['help'])
async def user_Help(message: types.Message):
    await message.reply("Бог вам в помощь!")

@dp.message_handler(commands=['song'])
async def user_Help(message: types.Message):
    await message.reply("МАМАААААААА УУУУУУУ!")

@dp.message_handler(commands=['song2'])
async def user_Help(message: types.Message):
    await message.reply("Half past twelve\nAnd I'm watching the late show in my flat all alone\nHow I hate to spend the evening on my own")
    await message.reply("There's not a soul out there\nNo one to hear my prayer!")
    await message.reply("Gimme gimme gimme a man after midnight\nWon't somebody help me chase these shadows away\nGimme gimme gimme a man after midnight\nTake me through the darkness to the break of the day!")

async def confirmation(message: types.Message,parameter:int):
    global choosed
    choosed=parameter
    await message.reply("Вы выбрали тип забега:\n" + str(message.text))
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(*buttons_1)
    await message.answer("Всё верно?", reply_markup=keyboard)

@dp.message_handler(filters.Text(equals='1. Забег за пивом'))
async def confirmation_1(message:types.Message):
    await confirmation(message,0)

@dp.message_handler(filters.Text(equals="2. Забег без правил"))
async def confirmation_2(message: types.Message):
    await confirmation(message,1)

@dp.message_handler(filters.Text(equals="3. Классический"))
async def confirmation_3(message: types.Message):
    await confirmation(message,2)

@dp.message_handler(filters.Text(equals="4. Льготный"))
async def confirmation_4(message: types.Message):
    await confirmation(message,3)

@dp.message_handler(filters.Text(equals="Да"))
async def data_Input(message: types.Message):
    await message.reply("Введите свои ФИО через пробел:\n пример заполнения: Кабанов Артемий Андреевич")

@dp.message_handler(filters.Text(equals="Нет"))
async def restart(message: types.Message):
    await message.reply("Ошиблись? Выберите снова!")
    global choosed
    choosed=-1

    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(*buttons)

    await message.answer("Просто выберите нужный забег:", reply_markup=keyboard)



@dp.message_handler(content_types=['text'])
async def data_To_Ex_Checker(message: types.Message):
    global user_data
    user_data=message.text.split()
    if len(user_data)==3 and choosed>=0:
        data_To_Ex()
        await message.reply("успешно зарегистрирован(а)!\n Ждём вас 31 Февраля 2044 года у станции метро Заречная")
    else:
        await  message.reply("Нарушен формат заполнения или не выбран тип забега!")


def data_To_Ex():
    counter_y = 2
    counter_x = 2
    while (True):
        cell = ws_list[choosed].cell(counter_y, counter_x)
        if cell.value is None:
            for ell in zip(user_data):
                cell.value = ''.join(ell)
                counter_x += 1
                cell = ws_list[choosed].cell(counter_y, counter_x)
            break
        counter_y += 1
    lwb.save(my_list)  # сохраняем и закрываем лист
    lwb.close()


if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)



